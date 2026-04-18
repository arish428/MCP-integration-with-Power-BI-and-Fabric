from pathlib import Path
import json
from openpyxl import Workbook

import compare_powerbi_models as model_compare


SOX_REMOVED_COLUMNS = [
    "my_row_id",
    "BK_CustomerId",
    "reviewIssue",
    "reviewIssueType",
    "reviewStatus",
    "reviewAssignee",
    "reviewCreator",
    "reviewCreated",
    "reviewUpdated",
    "attestIssueType",
    "attestSummary",
    "attestCreator",
    "attestAssignee",
    "attestCreated",
    "attestUpdated",
    "buId",
    "soxTemplateId",
    "Completed By User",
    "Completed By User Firstname",
    "Completed By User Lastname",
    "completedOn",
    "OU_Id",
    "president_username",
    "president_firstName",
    "president_lastName",
    "ceo_username",
    "ceo_firstName",
    "ceo_lastName",
    "pre_ceo_fullnames",
    "BK_TemplateInstancesId",
]


def autofit_columns(worksheet) -> None:
    for column in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        worksheet.column_dimensions[column[0].column_letter].width = min(width, 60)


def main() -> None:
    adomd_dir = Path(r"C:\Program Files\Microsoft.NET\ADOMD.NET\160")
    model_compare.ADOMD_DIR = adomd_dir
    model_compare.NATIVE_DIR = adomd_dir
    model_compare.configure_adomd()

    discovered = model_compare.discover_open_models()
    selected: dict[str, list[dict[str, object]]] = {}
    for row in discovered:
        pbix_path = row["pbix_path"]
        if pbix_path.endswith("SOX Report (prod).pbix") or pbix_path.endswith("SOX Report (stage).pbix"):
            selected[pbix_path] = model_compare.fetch_model_metadata(row["port"])["columns"]

    prod_path = next(path for path in selected if path.endswith("(prod).pbix"))
    stage_path = next(path for path in selected if path.endswith("(stage).pbix"))

    prod_bare_columns = {str(column["column"]) for column in selected[prod_path]}
    stage_bare_columns = {str(column["column"]) for column in selected[stage_path]}

    # Analyze against the provided list
    missing_in_stage = [col for col in SOX_REMOVED_COLUMNS if col in prod_bare_columns and col not in stage_bare_columns]
    present_in_both = [col for col in SOX_REMOVED_COLUMNS if col in prod_bare_columns and col in stage_bare_columns]
    not_in_prod = [col for col in SOX_REMOVED_COLUMNS if col not in prod_bare_columns]

    output_path = Path(__file__).resolve().parent / "sox_removed_columns_analysis.xlsx"

    workbook = Workbook()
    
    # Sheet 1: Missing in Stage (should be removed from stage)
    sheet_missing = workbook.active
    sheet_missing.title = "Missing in Stage"
    sheet_missing.append(["Column", "Exists in Prod", "Exists in Stage", "Status"])
    for col in missing_in_stage:
        sheet_missing.append([col, "Yes", "No", "Should be removed from stage"])

    # Sheet 2: Present in Both
    sheet_both = workbook.create_sheet("Present in Both")
    sheet_both.append(["Column", "Exists in Prod", "Exists in Stage", "Status"])
    for col in present_in_both:
        sheet_both.append([col, "Yes", "Yes", "Already present in both"])

    # Sheet 3: Not in Prod
    sheet_not_prod = workbook.create_sheet("Not in Prod")
    sheet_not_prod.append(["Column", "Exists in Prod", "Exists in Stage", "Status"])
    for col in not_in_prod:
        sheet_not_prod.append([col, "No", "Unknown", "Does not exist in prod"])

    # Sheet 4: Summary
    sheet_summary = workbook.create_sheet("Summary")
    sheet_summary.append(["Category", "Count"])
    sheet_summary.append(["Missing in Stage (should be removed)", len(missing_in_stage)])
    sheet_summary.append(["Present in Both", len(present_in_both)])
    sheet_summary.append(["Not in Prod", len(not_in_prod)])
    sheet_summary.append(["Total Columns Checked", len(SOX_REMOVED_COLUMNS)])

    for ws in workbook.sheetnames:
        autofit_columns(workbook[ws])

    workbook.save(output_path)
    print(f"Excel file created: {output_path}")
    print(json.dumps(
        {
            "missing_in_stage": missing_in_stage,
            "present_in_both": present_in_both,
            "not_in_prod": not_in_prod,
        },
        indent=2,
    ))


if __name__ == "__main__":
    main()
