from pathlib import Path
import json
from openpyxl import Workbook

import compare_powerbi_models as model_compare


REMOVED_COLUMNS_BY_TABLE = {
    "Exam Table": [
        "my_row_id",
        "updated",
        "numwatchers",
        "Label",
        "IssueEntityAgencyName",
        "IssueSourceName",
        "Owner",
        "Planned Completion Date",
        "Organizational Hierarchy Text",
        "Site Text",
        "Exam Closed",
    ],
    "SubTask Table": [
        "my_row_id",
        "PK_IssueId",
        "BK_IssueId",
        "issuetype",
        "reporter",
        "resolutiondate",
        "creator",
        "NumOfWatchers",
        "Occurrences",
        "End date",
        "Group Assignee",
        "Assignee Type",
        "Recur every",
        "Recur every week on",
        "Ends",
        "siteName",
        "Recurrence",
        "Watchers",
        "Sub Type",
        "Business Unit Id",
        "Business Unit Name",
        "Task_BK_IssueId",
        "CustomerId",
        "Priority",
    ],
    "Issues Table": [
        "my_row_id",
        "PK_IssueId",
        "assignee",
        "resolutiondate",
        "creator",
        "numwatchers",
        "IssueSource",
        "AgencyOrEntity",
        "TypeOfIssue",
        "IssueIdentificationDate",
        "TiedToAuditExamDescription",
        "Parent BU ID",
        "Parent BU",
        "ResponsibleDepartment",
        "ResolutionDueDate",
        "Severity In Number",
        "RootCause",
        "RootCauseDescription",
        "SubmitterName",
        "RepeatFinding",
        "IssueResolutionDate",
        "FinancialReportingImpact",
        "Priority",
        "ProjectNumber",
        "AssigneeType",
        "GroupAssignee",
        "Comments",
        "Subject Area",
        "Action Plan Response",
        "Action Plan IssueStatus",
        "Action Plan Created",
        "Action Plan DueDate",
        "Action Plan Revised Due Date",
        "Action Plan Last Status Update",
        "Action Plan Last Status Update Date",
        "Organizational Hierarchy/Business Unit/Branch",
        "Risk Category",
        "Assignee Full Name",
        "Risk",
        "Risk Business Unit",
        "SOX Control Impact",
        "Validated By 1 Text",
        "Validated Date 1",
        "Validated By 2 Text",
        "Validated Date 2",
        "Validated By 3 Text",
        "Validated Date 3",
        "BK_RiskRegisterItemId",
        "Lvl1",
        "Date Closed",
        "Actual Loss",
        "Potential Loss",
        "JoiningKey",
        "PK_RiskRegisterId",
        "BK_RiskRegisterId",
        "Status_Update",
        "Site Text",
        "Issue Workflow",
    ],
    "Exam Custom Field Table": [
        "customerCode",
        "BK_CustomerId",
        "riskId",
    ],
}


def autofit_columns(worksheet) -> None:
    for column in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        worksheet.column_dimensions[column[0].column_letter].width = min(width, 60)


def main() -> None:
    adomd_dir = Path(r"C:\Program Files\Microsoft.NET\ADOMD.NET\160")
    model_compare.ADOMD_DIR = adomd_dir
    model_compare.NATIVE_DIR = adomd_dir
    model_compare.configure_adomd()

    discovered = model_compare.discover_open_models()
    selected: dict[str, list[dict[str, object]]] = {}
    for row in discovered:
        pbix_path = row["pbix_path"]
        if pbix_path.endswith("Exams with Requests List (prod).pbix") or pbix_path.endswith(
            "Exams with Requests List (stage).pbix"
        ):
            metadata = model_compare.fetch_model_metadata(row["port"])
            selected[pbix_path] = metadata["columns"]

    prod_path = next(path for path in selected if path.endswith("(prod).pbix"))
    stage_path = next(path for path in selected if path.endswith("(stage).pbix"))

    prod_bare_columns = {str(column["column"]) for column in selected[prod_path]}
    stage_bare_columns = {str(column["column"]) for column in selected[stage_path]}

    output_path = Path(__file__).resolve().parent / "exams_removed_columns_analysis.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_data = []
    all_table_results = {}

    for table_name, columns_list in REMOVED_COLUMNS_BY_TABLE.items():
        missing_in_stage = [col for col in columns_list if col in prod_bare_columns and col not in stage_bare_columns]
        present_in_both = [col for col in columns_list if col in prod_bare_columns and col in stage_bare_columns]
        not_in_prod = [col for col in columns_list if col not in prod_bare_columns]

        all_table_results[table_name] = {
            "missing_in_stage": missing_in_stage,
            "present_in_both": present_in_both,
            "not_in_prod": not_in_prod,
        }

        summary_data.append({
            "table": table_name,
            "total_checked": len(columns_list),
            "missing_in_stage": len(missing_in_stage),
            "present_in_both": len(present_in_both),
            "not_in_prod": len(not_in_prod),
        })

        # Create sheet for each table
        sheet = workbook.create_sheet(table_name)
        sheet.append(["Column", "Exists in Prod", "Exists in Stage", "Status"])
        for col in missing_in_stage:
            sheet.append([col, "Yes", "No", "Missing in Stage"])
        for col in present_in_both:
            sheet.append([col, "Yes", "Yes", "Present in Both"])
        for col in not_in_prod:
            sheet.append([col, "No", "Unknown", "Not in Prod"])
        autofit_columns(sheet)

    # Create Summary sheet
    sheet_summary = workbook.create_sheet("Summary", 0)
    sheet_summary.append(["Table Name", "Total Columns Checked", "Missing in Stage", "Present in Both", "Not in Prod"])
    for item in summary_data:
        sheet_summary.append([
            item["table"],
            item["total_checked"],
            item["missing_in_stage"],
            item["present_in_both"],
            item["not_in_prod"],
        ])
    autofit_columns(sheet_summary)

    workbook.save(output_path)
    print(f"Excel file created: {output_path}")
    print(json.dumps(all_table_results, indent=2))


if __name__ == "__main__":
    main()
